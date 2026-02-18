"""Report generation module"""
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


class ReportGenerator:
    """Generate monthly reports"""
    
    def __init__(self, excel_mgr):
        self.excel_mgr = excel_mgr
    
    def generate_report(self, centre_name="Centre Principal"):
        """Generate formatted monthly report"""
        try:
            # Load insertions
            insertions = self.excel_mgr.load_insertions()
            
            if not insertions:
                return False, "No insertions to export"
            
            # Create report filename
            current_month = datetime.now().strftime("%B_%Y")
            report_file = os.path.join(
                self.excel_mgr.base_path,
                f'rapport_{current_month}.xlsx'
            )
            
            # Load the rapport file as template
            wb = load_workbook(self.excel_mgr.rapport_file)
            ws = wb['Insertions']
            
            # Add header section
            ws.insert_rows(1, 3)
            
            # Header row 1
            ws.merge_cells('A1:N1')
            cell = ws['A1']
            cell.value = "Direction SAV Géant Froid"
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Header row 2
            ws.merge_cells('A2:N2')
            cell = ws['A2']
            cell.value = f"Centre SAV {centre_name}"
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Header row 3
            ws.merge_cells('A3:N3')
            cell = ws['A3']
            cell.value = f"Rapport Mois {datetime.now().strftime('%B %Y')}"
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Style the data headers (now row 4)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders to all data
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
            
            # Adjust column widths
            column_widths = {
                'A': 5,   # #
                'B': 20,  # Client
                'C': 20,  # Produit
                'D': 20,  # Type de produit
                'E': 15,  # N° de série
                'F': 15,  # Garantie
                'G': 12,  # Date produit
                'H': 20,  # Panne
                'I': 25,  # Réparation effectuée
                'J': 20,  # PDR consommée
                'K': 15,  # Statut
                'L': 20,  # Centre
                'M': 12,  # Date réception
                'N': 12   # Date réparation
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Enable auto-filter on header row
            ws.auto_filter.ref = f'A4:N{ws.max_row}'
            
            # Save report
            wb.save(report_file)
            wb.close()
            
            return True, f"Report generated successfully: {report_file}"
        
        except Exception as e:
            return False, f"Failed to generate report: {str(e)}"
