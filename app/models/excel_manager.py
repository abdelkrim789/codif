"""Excel file manager for reading and writing codification data"""
import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


class ExcelManager:
    """Manages reading and writing to Excel files"""
    
    def __init__(self, base_path='data'):
        self.base_path = base_path
        self.reference_file = os.path.join(base_path, 'codification_reference.xlsx')
        self.rapport_file = os.path.join(base_path, 'rapport_insertions.xlsx')
        self.archives_dir = os.path.join(base_path, 'archives')
    
    def create_reference_file(self):
        """Create the codification reference Excel file with all required sheets"""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        wb.create_sheet('Familles')
        wb.create_sheet('Produits')
        wb.create_sheet('Models')
        wb.create_sheet('Pannes')
        wb.create_sheet('Causes')
        wb.create_sheet('Solutions')
        wb.create_sheet('PDR')
        wb.create_sheet('Centres')
        wb.create_sheet('Users')
        
        # Setup Familles sheet
        ws_familles = wb['Familles']
        ws_familles.append(['ID', 'Famille'])
        
        # Setup Produits sheet
        ws_produits = wb['Produits']
        ws_produits.append(['ID', 'Famille_ID', 'Produit'])
        
        # Setup Models sheet
        ws_models = wb['Models']
        ws_models.append(['ID', 'Produit_ID', 'Model'])
        
        # Setup Pannes sheet
        ws_pannes = wb['Pannes']
        ws_pannes.append(['ID', 'Code', 'Produit_ID', 'Panne'])
        
        # Setup Causes sheet
        ws_causes = wb['Causes']
        ws_causes.append(['ID', 'Code', 'Panne_ID', 'Cause'])
        
        # Setup Solutions sheet
        ws_solutions = wb['Solutions']
        ws_solutions.append(['ID', 'Code', 'Cause_ID', 'Solution'])
        
        # Setup PDR sheet
        ws_pdr = wb['PDR']
        ws_pdr.append(['ID', 'Code', 'PDR'])
        
        # Setup Centres sheet
        ws_centres = wb['Centres']
        ws_centres.append(['ID', 'Centre'])
        
        # Setup Agents Agréés sheet
        wb.create_sheet('Agents')
        ws_agents = wb['Agents']
        ws_agents.append(['ID', 'Nom_Prenom', 'Telephone', 'Wilaya', 'Adresse'])
        
        # Setup Users sheet
        ws_users = wb['Users']
        ws_users.append(['ID', 'Username', 'Password', 'Role'])
        # Add default admin account
        ws_users.append([1, 'admin', 'admin123', 'admin'])
        
        wb.save(self.reference_file)
        return wb
    
    def create_rapport_file(self):
        """Create the rapport insertions Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Insertions'
        
        # Setup headers
        headers = ['#', 'Client', 'Produit', 'Type de produit', 'N° de série', 
                   'Garantie', 'Date produit', 'Panne', 'Réparation effectuée', 
                   'PDR consommée', 'Statut', 'Centre', 'Date réception', 'Date réparation']
        
        ws.append(headers)
        
        # Style the headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Enable auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(self.rapport_file)
        return wb
    
    def load_reference_data(self):
        """Load all reference data from Excel file"""
        if not os.path.exists(self.reference_file):
            return None
        
        wb = load_workbook(self.reference_file)
        data = {}
        
        # Load Familles
        data['familles'] = []
        ws = wb['Familles']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data['familles'].append({'id': row[0], 'famille': row[1]})
        
        # Load Produits
        data['produits'] = []
        ws = wb['Produits']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data['produits'].append({'id': row[0], 'famille_id': row[1], 'produit': row[2]})
        
        # Load Models
        data['models'] = []
        ws = wb['Models']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data['models'].append({'id': row[0], 'produit_id': row[1], 'model': row[2]})
        
        # Load Pannes
        data['pannes'] = []
        ws = wb['Pannes']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data['pannes'].append({'id': row[0], 'code': row[1], 'produit_id': row[2], 'panne': row[3]})
        
        # Load Causes
        data['causes'] = []
        ws = wb['Causes']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data['causes'].append({'id': row[0], 'code': row[1], 'panne_id': row[2], 'cause': row[3]})
        
        # Load Solutions
        data['solutions'] = []
        ws = wb['Solutions']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data['solutions'].append({'id': row[0], 'code': row[1], 'cause_id': row[2], 'solution': row[3]})
        
        # Load PDR
        data['pdr'] = []
        ws = wb['PDR']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data['pdr'].append({'id': row[0], 'code': row[1], 'pdr': row[2]})
        
        # Load Centres
        data['centres'] = []
        ws = wb['Centres']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data['centres'].append({'id': row[0], 'centre': row[1]})
        
        # Load Agents Agréés
        data['agents'] = []
        if 'Agents' in wb.sheetnames:
            ws = wb['Agents']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    data['agents'].append({
                        'id': row[0],
                        'nom_prenom': row[1],
                        'telephone': row[2] if len(row) > 2 else '',
                        'wilaya': row[3] if len(row) > 3 else '',
                        'adresse': row[4] if len(row) > 4 else ''
                    })
        
        # Load Users
        data['users'] = []
        ws = wb['Users']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data['users'].append({'id': row[0], 'username': row[1], 'password': row[2], 'role': row[3]})
        
        wb.close()
        return data
    
    def save_reference_data(self, data):
        """Save reference data to Excel file"""
        wb = load_workbook(self.reference_file)
        
        # Save Familles
        ws = wb['Familles']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('familles', []):
            ws.append([item['id'], item['famille']])
        
        # Save Produits
        ws = wb['Produits']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('produits', []):
            ws.append([item['id'], item['famille_id'], item['produit']])
        
        # Save Models
        ws = wb['Models']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('models', []):
            ws.append([item['id'], item['produit_id'], item['model']])
        
        # Save Pannes
        ws = wb['Pannes']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('pannes', []):
            ws.append([item['id'], item['code'], item['produit_id'], item['panne']])
        
        # Save Causes
        ws = wb['Causes']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('causes', []):
            ws.append([item['id'], item['code'], item['panne_id'], item['cause']])
        
        # Save Solutions
        ws = wb['Solutions']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('solutions', []):
            ws.append([item['id'], item['code'], item['cause_id'], item['solution']])
        
        # Save PDR
        ws = wb['PDR']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('pdr', []):
            ws.append([item['id'], item['code'], item['pdr']])
        
        # Save Centres
        ws = wb['Centres']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('centres', []):
            ws.append([item['id'], item['centre']])
        
        # Save Agents Agréés
        if 'Agents' not in wb.sheetnames:
            wb.create_sheet('Agents')
            wb['Agents'].append(['ID', 'Nom_Prenom', 'Telephone', 'Wilaya', 'Adresse'])
        ws = wb['Agents']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('agents', []):
            ws.append([item['id'], item['nom_prenom'], item['telephone'], item['wilaya'], item['adresse']])
        
        # Save Users
        ws = wb['Users']
        ws.delete_rows(2, ws.max_row)
        for item in data.get('users', []):
            ws.append([item['id'], item['username'], item['password'], item['role']])
        
        wb.save(self.reference_file)
        wb.close()
    
    def load_insertions(self):
        """Load all insertions from rapport file"""
        if not os.path.exists(self.rapport_file):
            return []
        
        wb = load_workbook(self.rapport_file)
        ws = wb['Insertions']
        
        insertions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                insertions.append({
                    'num': row[0],
                    'client': row[1],
                    'produit': row[2],
                    'type_produit': row[3],
                    'num_serie': row[4],
                    'garantie': row[5],
                    'date_produit': row[6],
                    'panne': row[7],
                    'reparation': row[8],
                    'pdr': row[9],
                    'statut': row[10],
                    'centre': row[11],
                    'date_reception': row[12],
                    'date_reparation': row[13]
                })
        
        wb.close()
        return insertions
    
    def add_insertion(self, insertion_data):
        """Add a new insertion to the rapport file"""
        wb = load_workbook(self.rapport_file)
        ws = wb['Insertions']
        
        # Get next row number
        next_num = ws.max_row
        
        ws.append([
            next_num,
            insertion_data.get('client'),
            insertion_data.get('produit'),
            insertion_data.get('type_produit'),
            insertion_data.get('num_serie'),
            insertion_data.get('garantie'),
            insertion_data.get('date_produit'),
            insertion_data.get('panne'),
            insertion_data.get('reparation'),
            insertion_data.get('pdr'),
            insertion_data.get('statut'),
            insertion_data.get('centre'),
            insertion_data.get('date_reception'),
            insertion_data.get('date_reparation')
        ])
        
        wb.save(self.rapport_file)
        wb.close()
    
    def import_reference_from_excel(self, file_path):
        """Import reference data from an Excel file with matching structure.
        
        The Excel file should have sheets named: Familles, Produits, Models,
        Pannes, Causes, Solutions, PDR, Centres, Agents (optional).
        Data is merged (appended) into the existing reference data with new IDs.
        """
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}"
        
        try:
            wb_import = load_workbook(file_path, data_only=True)
        except Exception as e:
            return False, f"Cannot open Excel file: {str(e)}"
        
        # Load current data
        current_data = self.load_reference_data()
        if not current_data:
            return False, "Cannot load current reference data"
        
        imported_counts = {}
        
        # Map of sheet -> (data_key, columns)
        sheet_map = {
            'Familles': ('familles', ['id', 'famille']),
            'Produits': ('produits', ['id', 'famille_id', 'produit']),
            'Models': ('models', ['id', 'produit_id', 'model']),
            'Pannes': ('pannes', ['id', 'code', 'produit_id', 'panne']),
            'Causes': ('causes', ['id', 'code', 'panne_id', 'cause']),
            'Solutions': ('solutions', ['id', 'code', 'cause_id', 'solution']),
            'PDR': ('pdr', ['id', 'code', 'pdr']),
            'Centres': ('centres', ['id', 'centre']),
            'Agents': ('agents', ['id', 'nom_prenom', 'telephone', 'wilaya', 'adresse']),
        }
        
        for sheet_name, (data_key, columns) in sheet_map.items():
            if sheet_name not in wb_import.sheetnames:
                continue
            
            ws = wb_import[sheet_name]
            next_id = max([item['id'] for item in current_data.get(data_key, [])], default=0) + 1
            count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                item = {}
                for i, col_name in enumerate(columns):
                    if i < len(row):
                        val = row[i]
                        item[col_name] = val if val is not None else ''
                    else:
                        item[col_name] = ''
                
                # Assign new ID
                item['id'] = next_id
                next_id += 1
                count += 1
                
                current_data[data_key].append(item)
            
            if count > 0:
                imported_counts[sheet_name] = count
        
        wb_import.close()
        
        # Save merged data
        self.save_reference_data(current_data)
        
        if not imported_counts:
            return False, "No data found in the Excel file to import"
        
        summary = ", ".join([f"{v} {k}" for k, v in imported_counts.items()])
        return True, f"Successfully imported: {summary}"
    
    def import_agents_from_excel(self, file_path):
        """Import agents agréés from an Excel file.
        
        The Excel file should have columns: Nom_Prenom, Telephone, Wilaya, Adresse
        (first row is header).
        """
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}"
        
        try:
            wb_import = load_workbook(file_path, data_only=True)
        except Exception as e:
            return False, f"Cannot open Excel file: {str(e)}"
        
        ws = wb_import.active
        
        current_data = self.load_reference_data()
        if not current_data:
            return False, "Cannot load current reference data"
        
        next_id = max([a['id'] for a in current_data.get('agents', [])], default=0) + 1
        count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            agent = {
                'id': next_id,
                'nom_prenom': str(row[0]) if row[0] else '',
                'telephone': str(row[1]) if len(row) > 1 and row[1] else '',
                'wilaya': str(row[2]) if len(row) > 2 and row[2] else '',
                'adresse': str(row[3]) if len(row) > 3 and row[3] else ''
            }
            current_data['agents'].append(agent)
            next_id += 1
            count += 1
        
        wb_import.close()
        
        if count == 0:
            return False, "No agent data found in the file"
        
        self.save_reference_data(current_data)
        return True, f"Successfully imported {count} agents agréés"
    
    # --- Archive methods ---
    
    def ensure_archives_dir(self):
        """Create archives directory if it doesn't exist"""
        os.makedirs(self.archives_dir, exist_ok=True)
    
    def import_archive(self, file_path):
        """Import a previous monthly report into the archives directory"""
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}"
        
        self.ensure_archives_dir()
        
        filename = os.path.basename(file_path)
        dest = os.path.join(self.archives_dir, filename)
        
        # Avoid overwriting
        if os.path.exists(dest):
            base, ext = os.path.splitext(filename)
            counter = 1
            while os.path.exists(dest) and counter < 1000:
                dest = os.path.join(self.archives_dir, f"{base}_{counter}{ext}")
                counter += 1
            if os.path.exists(dest):
                return False, "Too many copies of this file already archived"
        
        try:
            shutil.copy2(file_path, dest)
            return True, f"Archived: {os.path.basename(dest)}"
        except Exception as e:
            return False, f"Failed to archive: {str(e)}"
    
    def list_archives(self):
        """List all archived reports"""
        self.ensure_archives_dir()
        archives = []
        for f in sorted(os.listdir(self.archives_dir)):
            if f.endswith(('.xlsx', '.xls')):
                full_path = os.path.join(self.archives_dir, f)
                size = os.path.getsize(full_path)
                mtime = datetime.fromtimestamp(os.path.getmtime(full_path))
                archives.append({
                    'filename': f,
                    'path': full_path,
                    'size': size,
                    'date': mtime.strftime('%Y-%m-%d %H:%M')
                })
        return archives
    
    def load_archive_data(self, archive_path):
        """Load insertion data from an archived report file"""
        if not os.path.exists(archive_path):
            return []
        
        try:
            wb = load_workbook(archive_path, data_only=True)
        except Exception:
            return []
        
        # Try to find the data sheet
        ws = None
        if 'Insertions' in wb.sheetnames:
            ws = wb['Insertions']
        else:
            ws = wb.active
        
        insertions = []
        # Skip header rows (may have merged header rows at top)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                # Try to detect if this is a header row
                first_val = str(row[0]).strip()
                if first_val in ('#', 'Direction', 'Centre', 'Rapport', ''):
                    continue
                insertions.append({
                    'num': row[0] if row[0] else '',
                    'client': row[1] if len(row) > 1 and row[1] else '',
                    'produit': row[2] if len(row) > 2 and row[2] else '',
                    'type_produit': row[3] if len(row) > 3 and row[3] else '',
                    'num_serie': row[4] if len(row) > 4 and row[4] else '',
                    'garantie': row[5] if len(row) > 5 and row[5] else '',
                    'date_produit': row[6] if len(row) > 6 and row[6] else '',
                    'panne': row[7] if len(row) > 7 and row[7] else '',
                    'reparation': row[8] if len(row) > 8 and row[8] else '',
                    'pdr': row[9] if len(row) > 9 and row[9] else '',
                    'statut': row[10] if len(row) > 10 and row[10] else '',
                    'centre': row[11] if len(row) > 11 and row[11] else '',
                    'date_reception': row[12] if len(row) > 12 and row[12] else '',
                    'date_reparation': row[13] if len(row) > 13 and row[13] else ''
                })
        
        wb.close()
        return insertions
